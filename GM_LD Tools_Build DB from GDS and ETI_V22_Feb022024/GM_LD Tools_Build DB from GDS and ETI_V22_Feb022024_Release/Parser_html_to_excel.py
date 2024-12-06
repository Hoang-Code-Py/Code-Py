from ast import Break
from distutils import command
from pickle import TRUE
from bs4 import BeautifulSoup
import openpyxl
import re
import pandas as pd
from sympy import symbols, Eq, solve
import ast

# Function:
def Mapping_a_x_plus_b(keymap, Sheet_ax_plus_b_Index_TableID):
    if keymap in Sheet_ax_plus_b_Index_TableID.index:
        dfff = Sheet_ax_plus_b_Index_TableID.loc[keymap]
        if len(dfff) == 23:
            # Series:
            item = dfff.get(key="ItemID")
        else:
            item = dfff.iat[0, 0]
    else:
        item = "DB missing"
    return item


def Mapping_a_and_b(Key_ab, Sheet_a_and_b_Index_TableID):
    if Key_ab in Sheet_a_and_b_Index_TableID.index:
        dfff = Sheet_a_and_b_Index_TableID.loc[Key_ab]
        if len(dfff) == 16:
            # Series:
            item = dfff.get(key="ItemID")
        else:
            item = dfff.iat[0, 0]
    else:
        item = "DB missing"
    return item


def support_GDS(Key_check, Sheet_GDS):
    Var_exist = False
    Group = 'Need Check'
    for i in range(0, len(Sheet_GDS)):
        data_row = Sheet_GDS.iat[i, 3]
        if Key_check == data_row:
            Var_exist = True
            Group = Sheet_GDS.iat[i, 0]
            Break
    return Var_exist, Group


def xuliPID_to_command(byte3 , PID, byte_end):
    # ex: '$402D' => '04 2C FE 40 2D'
    command = ""
    try:
        byte1 = PID[1:3]
        byte2 = PID[-2:]
        command = command + byte3 + byte1 + ' ' + byte2 + byte_end
    except:
        pass
    return command


def main_Liss(List_Main_TableID, List_yourstring, Sheet_TableID_Index_TableID):

    if ["No Table ID in DB"] in List_Main_TableID:
        Return_value = "No Table ID in DB"
    else:
        VDSF = set.intersection(*[set(v) for v in List_Main_TableID])
        if VDSF == set():
            Return_value = "No Table ID in DB"
        else:
            Final = list(VDSF)
            Return_value_OK = ""
            Return_value_Fail = ""
            for IDCHECK in Final:
                LALALAL = Sheet_TableID_Index_TableID.loc[IDCHECK, :]
                print(LALALAL)
                print("------", LALALAL.iat[len(LALALAL)-1, 2])
                checkdefault = LALALAL.iat[len(LALALAL)-1, 2]
                # print(len(Sheet_TableID_Index_TableID.loc[IDCHECK,:]))
                if checkdefault == "DEFAULT":
                    lencompare = len(LALALAL)-1
                    print("len(List_yourstring)) DEFAULT",
                          len(List_yourstring))
                    if lencompare == len(List_yourstring):
                        Return_value_OK = IDCHECK
                    else:
                        Return_value_Fail = "No Table ID in DB"
                else:
                    lencompare = len(LALALAL)
                    print("len(List_yourstring)) NO DEFAULT",
                          len(List_yourstring))
                    if lencompare == len(List_yourstring):
                        Return_value_OK = IDCHECK
                    else:
                        Return_value_Fail = "No Table ID in DB"
            print("Return_value_OK : ", Return_value_OK)
            print("Return_value_Fail :", Return_value_Fail)
            if len(Return_value_OK) == 0:
                Return_value = Return_value_Fail
            else:
                Return_value = Return_value_OK
    return(Return_value)

def convert_format(input_list):
    output_list = []
    
    table_id = input_list[0]
    for item in input_list[1:]:
        parts = item.split(':')
        if parts[0] == 'DEFAULT':
            output_item = [table_id, 'DEFAULT', 'DEFAULT', 'N/A', str(parts[1])]
            output_list.append(output_item)
        else:
            Hex_1 = hex(int(parts[0]))
            Hex_Val = ''
            if len(Hex_1) == 3: #0x1
                Hex_Val = "0x0" + Hex_1.split('x')[1].upper()
            elif len(Hex_1) == 4: #0xff
                Hex_Val = "0x" + Hex_1.split('x')[1].upper()
            elif '0x' in Hex_1:
                if len(Hex_1) == 5: #case 2byte: 0x100 => Hex_1 = 100
                    Hex_Val = "0x0" + Hex_1.split('x')[1].upper()
                elif len(Hex_1) == 6: #case 2byte: 0x1100 => Hex_1 = 1100
                    Hex_Val = "0x" + Hex_1.split('x')[1].upper()
                elif len(Hex_1) == 7: #case 3byte: 0x10000 => 0x010000
                    Hex_Val = "0x0" + Hex_1.split('x')[1].upper()
                elif len(Hex_1) == 8: #case 3byte: 0x110000 => 0x110000
                    Hex_Val = "0x" + Hex_1.split('x')[1].upper()
                elif len(Hex_1) == 9: #case 4byte: 0x2020000=>  0x02020000
                    Hex_Val = "0x0" + Hex_1.split('x')[1].upper()
                else:
                    Hex_Val = "HEX VALUE: " + str(Hex_1)
            else:
                raise KeyError
            
            if len(Hex_Val) == 6:
                output_item = [table_id, '0xFFFF', str(Hex_Val), 'N/A', parts[1]]
                output_list.append(output_item)
            elif len(Hex_Val) == 8:
                output_item = [table_id, '0xFFFFFF', str(Hex_Val), 'N/A', parts[1]]
                output_list.append(output_item)
            elif len(Hex_Val) == 10:
                output_item = [table_id, '0xFFFFFFFF', str(Hex_Val), 'N/A', parts[1]]
                output_list.append(output_item)                
            else:
                output_item = [table_id, '0xFF', str(Hex_Val), 'N/A', parts[1]]
                output_list.append(output_item)        
    return output_list

def split_Table(String_x, LIST_KEYYYYMAPPPING, Sheet_TableID, Sheet_TableID_Index_TableID):
    # print(len(String_x))
    pos = 0
    List_yourstring = []
    for m in re.finditer('\n', String_x):
        print('enter found', m.start(), m.end())
        yourstring = String_x[pos:m.end()-1]
        x_string = yourstring.replace("\n", "")
        print("your string :", x_string)
        List_yourstring.append(x_string)
        pos = 0 + m.start()
    print("List_yourstring: ", List_yourstring)
    print("Len List_yourstring: ", len(List_yourstring))
    # Process for List_yourstring:
    List_Main_TableID = []

    checkValue = ""

    if len(List_yourstring) == 0:
        checkValue = "False Value"
    else:
        for i in List_yourstring:
            list_x = []
            index_i = i.find(":")
            Dec_val = i[0:index_i]
            Text_Val = i[index_i+1:]
            Hex_Val = ''
            # print("Dec_val :",Dec_val)
            # print("len(Dec_val)",type(Dec_val))
            # print("Text_Val :",Text_Val)
            # print("len(Text_Val)",len(Text_Val))
            if Dec_val != 'DEFAULT':
                Hex_1 = hex(int(Dec_val)).replace("0x", "").upper()
                if len(Hex_1) == 1:
                    Hex_Val = "0x0" + Hex_1
                elif len(Hex_1) == 3:
                    Hex_Val = "0x0" + Hex_1
                elif len(Hex_1) == 5:
                    Hex_Val = "0x0" + Hex_1     
                elif len(Hex_1) == 7:
                    Hex_Val = "0x0" + Hex_1                               
                else:
                    Hex_Val = "0x" + Hex_1
                # print("Hex_Val :",Hex_Val)
                # print("---------")
            KEYMAPPING = Hex_Val + Text_Val
            # print("============> ", type(Sheet_TableID.loc[KEYMAPPING,"TableID"]))
            # print("============> ", Sheet_TableID.loc[KEYMAPPING,"TableID"].iloc[0])
            if KEYMAPPING in LIST_KEYYYYMAPPPING:
                DF_LocKey = Sheet_TableID.loc[KEYMAPPING, "TableID"]
                print("DF_LocKey:", DF_LocKey)
                # print("Type :",type(DF_LocKey))
                if type(DF_LocKey) == str:
                    list_x.append(DF_LocKey)
                    List_Main_TableID.append(list_x)
                else:
                    for x in range(0, len(DF_LocKey)):
                        list_x.append(DF_LocKey.iloc[x])
                    List_Main_TableID.append(list_x)
            else:
                List_Main_TableID.append(["No Table ID in DB"])
        checkValue = "True Value"
    print("List_Main_TableID:", List_Main_TableID)
    print("checkValue : ", checkValue)
    if checkValue == "False Value":
        Your_Return_value = "No Table ID in DB"
    else:
        Your_Return_value = main_Liss(List_Main_TableID, List_yourstring, Sheet_TableID_Index_TableID)
    print("Your_Return_value :", Your_Return_value)
    return Your_Return_value, List_yourstring

def find_coefficients(expression):
    #'y=(1x+0)/0.0625'
    # Tách biểu thức thành các phần
    parts = expression.split('=')
    left = parts[0] 
    right = parts[1] #(1x+0)/0.0625'
    
    # Tìm hệ số b bằng cách lấy phần tử thứ 2 trong dấu ngoặc 
    heso = right.split('/') #['(1x+0)', '0.0625']

    # sau khi split theo dấu '+'
    heso[0] = heso[0].replace('(', '')
    heso[0] = heso[0].replace(')', '')

    coefs = heso[0].split('+') #['1x' , '0' ]

    #he so a:
    if coefs[0] == 'x':
        coefs[0] = coefs[0].replace('x','1')
    else:
        coefs[0] = coefs[0].replace('x','')

    a = float(coefs[0])/float(heso[1])

    b = 0
    if coefs[1] != '0':
        b = float(coefs[1])/float(heso[1])
    
   
    return a, b

def check_Table_ID(Sheet_Data):
    if Sheet_Data.empty:
        empty_df = pd.DataFrame()
        return Sheet_Data, empty_df
    else:
        # Danh sách mới của tên cột
        new_column_names = ['TableID', 'BIT_MASK', 'HEX_VAL', 'DEC_VAL', 'TABLE_TEXT']
        # Đặt tên mới cho các cột
        Sheet_Data.columns = new_column_names

        # print("Column names:", Sheet_Data.columns.tolist())
        Sheet_TableID_Index_TableID = Sheet_Data.copy()
        
        Sheet_TableID_Index_TableID .set_index(Sheet_TableID_Index_TableID.columns[0], inplace=True)

        # print('Sheet_Data: \n', Sheet_Data)
        # print('Sheet_TableID_Index_TableID: \n', Sheet_TableID_Index_TableID)

        List_TableID = []
        for i1 in range(0, len(Sheet_Data)):
            List_TableID.append(Sheet_Data['TableID'][i1])

        List_TableID = list(dict.fromkeys(List_TableID))
        key_list = []
        for table_id in List_TableID:
            # print("TableID đang check: ===> ", table_id)
            data_id = Sheet_TableID_Index_TableID.loc[table_id]
            # print("data_id ===> \n ", data_id)
            li = []
            # Converting dataframe to list
            li = data_id.values.tolist()
            li.sort()
            # print("li:\n", li)
            key_list.append(li)

        # list tong hop duplicate:
        list_tonghop = []

        print('# kiemtra')
        for tablle_id_2 in List_TableID:
            # print("TableID đang check: ===> ", tablle_id_2)
            data_id_2 = Sheet_TableID_Index_TableID.loc[tablle_id_2]
            value_list_check = []
            value_list_check = data_id_2.values.tolist()
            value_list_check.sort()
            list_duplicate = []
            for key_check in key_list:
                if value_list_check == key_check:
                    # print('===> Duplicate:')
                    index_tableid = key_list.index(key_check)
                    table_id_duplicate = List_TableID[index_tableid]
                    list_duplicate.append(tablle_id_2)
                    list_duplicate.append(table_id_duplicate)
            list_duplicate = list(dict.fromkeys(list_duplicate))
            # print("list_duplicate: \n", list_duplicate)
            list_tonghop.append(list_duplicate)
        # print("List tổng hợp: \n", list_tonghop)

        print("write value to txt file")
        list_keep = []
        for lis_xx in list_tonghop:
            if len(lis_xx) == 1:
                list_keep.append(lis_xx[0])
        # print("Keep value: \n", list_keep)
        string_in = ""
        for phan_tu in list_keep:
            string_in = string_in + '\n' + phan_tu
        # print("TableID khong bi duplicate: \n", string_in)

        print("Writte data to excel file")
        # nơi chứa data của 1 file html:
        DF_1 = pd.DataFrame({"ID need remove": ['x'], "ID Keep": ['x']})

        for lis_yy in list_tonghop:
            if len(lis_yy) == 2:
                if lis_yy[0] in list_keep:
                    Var_keep = lis_yy[0]
                    Var_Unkeep = lis_yy[1]
                else:
                    Var_keep = lis_yy[1]
                    Var_Unkeep = lis_yy[0]
                Sheet_Data['TableID'] = Sheet_Data['TableID'].apply(lambda x: Var_keep if x == Var_Unkeep else x)
                DF_2 = pd.DataFrame({"ID need remove": [Var_Unkeep], "ID Keep": [Var_keep]})
                frame = [DF_1, DF_2]
                DF_1 = pd.concat(frame)
        Sheet_Data = Sheet_Data.drop_duplicates(keep='first')
        print('Complete write')
        return Sheet_Data, DF_1


def CAN29_equation(expression):
    # y=(ax+b)/1
    match = re.search(r'y=\((.*?)\)/1', expression)
    if match:
        # print("Case Match")
        # result = match.group(1)
        # result_split = result.split('+')
        # a_value = result_split[0].replace('x', '')
        # if a_value == '':
        #     a_return = 1
        # else:
        #     a_return = a_value

        # b_return = result_split[1]
        a_return,b_return = find_coefficients(expression)
        return a_return , b_return
    else:
        try:
            print("Không tìm thấy phần ax+b trong biểu thức: ", expression )
            #'y=(1x+0)/0.0625'
            # y=(1x+230)/10
            a_return,b_return = find_coefficients(expression)
            return a_return , b_return
        except:
            raise KeyError


def run_parser_html2excel(Excel_PathFile, Sheet_GDS):
    wb = openpyxl.load_workbook('EXCEL.xlsx')
    ws = wb['Data']

    cHECK_BUG = False

    list_Table_Missing = []
    ID_Table_Missing = 2

    html_file = BeautifulSoup(open(Excel_PathFile), features="html.parser")

    Type_2262 = False
    found_string = html_file.find("td", string="DynamicIDs")
    if found_string == None:
        Type_2262 = True
    else:
        Type_2262 = False


    Sheet_TableID = pd.read_excel("EXCEL.xlsx", sheet_name="Table ID", index_col=0, header=0, convert_float=True, keep_default_na=False)
    Sheet_TableID_Noindex = pd.read_excel("EXCEL.xlsx", sheet_name="Table ID", header=0, convert_float=True, keep_default_na=False)
    Sheet_TableID_Index_TableID = pd.read_excel("EXCEL.xlsx", sheet_name="Table ID", index_col=1, header=0, convert_float=True, keep_default_na=False)

    Sheet_ax_plus_b_Index_TableID = pd.read_excel("EXCEL.xlsx", sheet_name="ax+b", index_col=22, header=0, convert_float=True, keep_default_na=False)
    print("Sheet_ax_plus_b_Index_TableID: \n", Sheet_ax_plus_b_Index_TableID)

    Sheet_a_and_b_Index_TableID = pd.read_excel("EXCEL.xlsx", sheet_name="axb", index_col=15, header=0, convert_float=True, keep_default_na=False)
    print("Sheet_a_and_b_Index_TableID: \n", Sheet_a_and_b_Index_TableID)

    Unit_path = r'Unit.xlsx'
    Unit_sheets = pd.read_excel(Unit_path, sheet_name='Sheet1', index_col=0)
    list_of_Unit = Unit_sheets.index.tolist()

    # print(Sheet_TableID)
    LIST_KEYYYYMAPPPING = []
    for abcsss in range(0, len(Sheet_TableID_Noindex)):
        # print("==============",Sheet_TableID_Noindex['Key Check GDS Sp'][abcsss])
        LIST_KEYYYYMAPPPING.append(Sheet_TableID_Noindex['Key Check GDS Sp'][abcsss])

    '''
    Nguyên Tắc Parser ETI:

    Cấu trúc:
    <body>
        <tr ALIGN = "Center">
            find(td).string -> PID $4077

            <pepid>
                find_all('td')=> cho vào 1 list => [PID name(0),Byte Position(1),Byte Size(2),xx(3),xx(4),xx(5),xx(6),Table(7)]
                    vào list[7] là 1 table: find_all('td'):
                        tạo 1 list chưa tất cả tag td: Nếu
                                                        + len(list) = 0 thì nó là một DID
                                                        + len(list) = 7 thì ("Đây là PID ax+b")
                                                            => phần tử trong list có ý nghĩa [xx(0), a(1),b(2),Max(3),Min(4),Unit(5),xx,xx,xx,..]
                                                        + còn lại  thì ("Đây là PID a&b")
                                                            => phần tử trong list có ý nghĩa [xx(0),xx(1),bit position(2),bit size(3),xx (4), table element(5->n))
                                                                => in các phần tử table ra
            </pepid>
            <pepid>
                PID name 2 ==> làm giống tag pepid trên
            </pepid>
            <pepid>
                PID name 3 ==> làm giống tag pepid trên
            </pepid>
            <pepid></pepid> => Nếu len của list chứa find_all('td) trong tag pepid không chứa gì cả = 0 -> không cần tới

        </tr>
    </body>
    '''
    # Kiểm tra CAN11 hay CAN29:
    List_Check_CAN = []
    List_check_CAN_1 = []
    for tag_h2 in html_file.find_all('h2'):
        List_Check_CAN.append(tag_h2)
    for tag_tr in List_Check_CAN[0].find_all('tr'):
        List_check_CAN_1.append(tag_tr)

    # print(List_check_CAN_1[-2])
    CANID = List_check_CAN_1[-2].find('td').string
    CANTYPE = ""
    if len(CANID) == 8:
        CANTYPE = 'CAN29Bit'
    elif len(CANID) == 4:
        CANTYPE = 'CAN11Bit'

    print("CANTYPE: ", CANTYPE)

    list1 = []
    for i in html_file.find_all('tr', attrs={"align": "Center"}):
        list1.append(i)
    exc_var = 2

    # Duyệt từng PID:
    for n in list1:
        # tim Name of PID:
        print("**********************************************************************")
        # print("PID:  ",n.find('td').string)
        print(".....................")
        lisstpep = []
        for peppp in n.find_all('pepid'):
            lisstpep.append(peppp)
        exc_var_2 = exc_var

        # Duyệt từng dòng của 1 PID:
        print('lEN lisstpep: ', len(lisstpep))
        if len(lisstpep) == 10:
            PID_string = n.find('td').string
            print("PID:  ", PID_string)
            
        else:
            for stt in lisstpep:
                listmini = []
                Command_AckCmd = 'N/A'
                Command_GetValueCmd = 'N/A'
                bytpos_write = ''
                for mini in stt.find_all('td'):
                    listmini.append(mini)

                if len(listmini) == 0:
                    print("fail r cha")
                else:
                    PID_string = n.find('td').string
                    print("PID:  ", PID_string)
                    if len(PID_string) <= 5:
                        command_write = xuliPID_to_command('04 2C FE ', n.find('td').string, '')
                        print("command_write:  ", command_write)

                        ws['A'+str(exc_var_2)] = n.find('td').string
                        if n.find('td').string == '$4527':
                            cHECK_BUG = True

                        Checktype03AA = n.find('td').string
                        bytepos = ""

                        if Type_2262 == True:
                            if CANTYPE == 'CAN11Bit':
                                if len(Checktype03AA) == 3:
                                    ws['D'+str(exc_var_2)] = Command_AckCmd
                                    Command_GetValueCmd = "03 AA 01 " + str(Checktype03AA[1:])
                                    ws['E'+str(exc_var_2)] = Command_GetValueCmd
                                    
                                    new_value2 = int(listmini[1].string) + 1
                                    bytepos = str(new_value2)
                                    ws['G'+str(exc_var_2)] = bytepos
                                    bytpos_write = bytepos
                                else: 
                                    ws['D'+str(exc_var_2)] = Command_AckCmd
                                    Command_GetValueCmd =  "03 22 " + str(command_write[9:])
                                    ws['E'+str(exc_var_2)] = Command_GetValueCmd
                                    new_value = 0
                                    if listmini[1].string[0] == '$':
                                        new_value = int(listmini[3].string) + 3
                                    else:
                                        new_value = int(listmini[1].string) + 3
                                    bytepos = str(new_value)
                                    ws['G'+str(exc_var_2)] = bytepos
                                    bytpos_write = bytepos
                            elif CANTYPE == 'CAN29Bit':
                                if len(Checktype03AA) != 3:
                                    ws['D'+str(exc_var_2)] = Command_AckCmd
                                    Command_GetValueCmd = "03 22 " + str(command_write[9:])
                                    ws['E'+str(exc_var_2)] = Command_GetValueCmd
                                    
                                    
                                    new_value = int(listmini[1].string) + 3
                                    bytepos = str(new_value)
                                    ws['G'+str(exc_var_2)] = bytepos
                                    bytpos_write = bytepos
                                else:
                                    raise KeyError
                        else:
                            if CANTYPE == 'CAN11Bit':
                                Command_AckCmd = command_write
                                ws['D'+str(exc_var_2)] = Command_AckCmd
                                Command_GetValueCmd = "03 AA 01 FE"
                                ws['E'+str(exc_var_2)] = Command_GetValueCmd
                                print("Byte Position:  ", listmini[1].string)
                                bytepos = listmini[1].string
                                ws['G'+str(exc_var_2)] = bytepos
                                bytpos_write = bytepos
                            elif CANTYPE == 'CAN29Bit':
                                #2C 01 F2 FE xx xx 01 04
                                sizePID = n.find_all('td')[1].string
                                number = int(sizePID)
                                hex_string = format(number, '02X')

                                command_write_29bit = xuliPID_to_command('2C 01 F2 FE ', n.find('td').string,' 01 ' + hex_string)
                                Command_AckCmd = command_write_29bit

                                ws['D'+str(exc_var_2)] = Command_AckCmd
                                Command_GetValueCmd = "03 2A 01 FE"
                                ws['E'+str(exc_var_2)] = Command_GetValueCmd
                                
                                print("Byte Position:  ", listmini[1].string)
                                bytepos = listmini[1].string
                                ws['G'+str(exc_var_2)] = 0
                                bytpos_write = '0'
                        
                        ws['F'+str(exc_var_2)] = "N/A"
                        print("PID Name:    ", listmini[0].string)
                        ws['B'+str(exc_var_2)] = listmini[0].string
                        print("ItemDescription:  ", listmini[0].string)
                        ws['C'+str(exc_var_2)] = listmini[0].string

                        bytesize_can29 = n.find_all('td')[1].string
                        if CANTYPE == 'CAN11Bit':
                            print("Byte Size:  ", listmini[2].string)
                            ws['H'+str(exc_var_2)] = listmini[2].string
                        else:
                            ws['H'+str(exc_var_2)] = bytesize_can29


                        print("=======>listmini:", listmini)
                        print("InputType:    ", listmini[3].string)
                        signout = ''
                        if listmini[3].string == 'Unsigned Integer':
                            signout = 'Unsigned'
                        elif listmini[3].string == 'Signed Integer':
                            signout = 'Signed'
                        else:
                            signout = listmini[3].string
                       
                        ws['M'+str(exc_var_2)] = signout

                        if CANTYPE == 'CAN11Bit':
                            lisss = []
                            # lisss chứa table của dang ax+b và a&b (cột readConverssion)
                            for var1 in listmini[7].find_all('td'):
                                lisss.append(var1)

                            if len(lisss) == 0:
                                print("=================================\n")
                                print("Đây là DID")
                                # vì DIC ở vị trí này là 1 control Mask không phải Table nên không có td
                                

                                SEV_positions = [index for index, item in enumerate(listmini) if 'SEV' == item.string]
                                Linear_positions = [index for index, item in enumerate(listmini) if 'Linear' == item.string]
                                ByteArray_positions = [index for index, item in enumerate(listmini) if 'Byte Array' == item.string]
                                # WIDTH_positions = [index for index, item in enumerate(listmini) if item.string is not None and (td_tag := item.find('td')) is not None and 'width' in td_tag.attrs]
                                WIDTH_positions = [index for index, item in enumerate(listmini) if (td_tag := BeautifulSoup(str(item), 'html.parser').find('td')) is not None and 'width' in td_tag.attrs]


                                if SEV_positions:
                                    print(f"Vị trí của SEV_positions: {SEV_positions}")
                                    Vitristart = SEV_positions[-1]
                                    Vitriend = WIDTH_positions[-1]
                                    BITPOS_VALUE = listmini[Vitristart+2].string
                                    BitSize_VALUE = listmini[Vitristart+3].string
                                    ws['I'+str(exc_var_2)] = BITPOS_VALUE
                                    ws['J'+str(exc_var_2)] = BitSize_VALUE
                                    ws['N'+str(exc_var_2)] = 'f(x)= x&a'
                                    ws['W'+str(exc_var_2)] = listmini[0].string + 'None'
                                    key_check_support_GDS = listmini[0].string + 'None'

                                    strrr = ""
                                    var_table = 1
                                   
                                    Default_text = 'DEFAULT:' + listmini[Vitristart+1].string

                                    lisssx = listmini[Vitristart+5:Vitriend]
                                    for xx in range(0, len(lisssx)):
                                        if lisssx[xx].string is not None:
                                            # print(lisss[xx].string)
                                            if var_table % 2 == 0:
                                                strrr = strrr + lisssx[xx].string + "\n"
                                            else:
                                                strrr = strrr + lisssx[xx].string + ":"
                                            var_table += 1
                                    print("---->>>:", strrr)
                                    

                                    TableID_text = ""
                                    if len(strrr) == 0:
                                        ws['O'+str(exc_var_2)] = "No Table ID in DB"
                                        TableID_text = "No Table ID in DB"
                                    else:
                                        TableID_text, Table_Mising = split_Table(strrr, LIST_KEYYYYMAPPPING, Sheet_TableID, Sheet_TableID_Index_TableID)
                                        if TableID_text == 'No Table ID in DB':
                                            ID_Table_Missing = ID_Table_Missing + 1
                                            ID_Table_Missing_str = 'TableID_Miss_00' + str(ID_Table_Missing)
                                            Table_Mising.append(Default_text)
                                            Table_Mising.insert(0, ID_Table_Missing_str)
                                            for convertaa in convert_format(Table_Mising):
                                                list_Table_Missing.append(convertaa)
                                            ws['O'+str(exc_var_2)] = ID_Table_Missing_str
                                        else:                               
                                            ws['O'+str(exc_var_2)] = TableID_text
                                    
                                
                                    Key_ab = listmini[0].string + Command_AckCmd + Command_GetValueCmd + bytpos_write + listmini[2].string + str(BITPOS_VALUE) + str(BitSize_VALUE) + TableID_text

                                    # print("Key_ab=   ", Key_ab)
                                    ws['Y'+str(exc_var_2)] = Key_ab

                                    Bien_kiemtra_GDS_2, Groupnamestring = support_GDS(key_check_support_GDS, Sheet_GDS)
                                    if Bien_kiemtra_GDS_2 == True:
                                        ItemID = Mapping_a_and_b(Key_ab, Sheet_a_and_b_Index_TableID)
                                    else:
                                        ItemID = 'GDS Not support'
                                    print("ItemID =   ", ItemID)

                                    ws['Z'+str(exc_var_2)] = ItemID
                                    ws['AA'+str(exc_var_2)] = Groupnamestring
                                    print("=========================\n")



                                elif Linear_positions:
                                    print(f"Vị trí của Linear_positions: {Linear_positions}")
                                    Vitristart = Linear_positions[-1]
                                    heso_a = listmini[Vitristart+1].string
                                    heso_b = listmini[Vitristart+2].string
                                    maxvalue = listmini[Vitristart+3].string
                                    minvalue = listmini[Vitristart+4].string
                                    if listmini[Vitristart+5].string  == '_':
                                        Unitfx = 'None'
                                        Unit_KEY = 'N/A'
                                    else:
                                        Unitfx = listmini[Vitristart+5].string 
                                        Unit_KEY = listmini[Vitristart+5].string 
                                    Floatvalue = listmini[Vitristart+6].string

                                    ws['N'+str(exc_var_2)] = 'f(x)= a*x+b'
                                    ws['P'+str(exc_var_2)] = heso_a
                                    ws['Q'+str(exc_var_2)] = heso_b
                                    ws['R'+str(exc_var_2)] = minvalue
                                    ws['S'+str(exc_var_2)] = maxvalue
                                    ws['T'+str(exc_var_2)] = Floatvalue
                                    ws['U'+str(exc_var_2)] = Unit_KEY

                                    Key_Check_GDS_Sp_FX = listmini[0].string + Unitfx        
                                    ws['W'+str(exc_var_2)] = Key_Check_GDS_Sp_FX
                                    Bien_kiemtra_GDS_1, Groupnamestring = support_GDS(Key_Check_GDS_Sp_FX, Sheet_GDS)
                                    
                                    Key_axcongb = listmini[0].string + Command_AckCmd + Command_GetValueCmd + bytepos + listmini[2].string + 'N/AN/A' + heso_a + heso_b + Unit_KEY
                                    ws['X'+str(exc_var_2)] = Key_axcongb
                                    if Bien_kiemtra_GDS_1 == True:
                                        ItemID = Mapping_a_x_plus_b(Key_axcongb, Sheet_ax_plus_b_Index_TableID)
                                    else:
                                        ItemID = 'GDS Not support'

                                    print("ItemID =   ", ItemID)
                                    ws['Z'+str(exc_var_2)] = ItemID
                                    ws['AA'+str(exc_var_2)] = Groupnamestring


                                elif ByteArray_positions:
                                    print(f"Vị trí của ByteArray_positions: {ByteArray_positions}")
                                else:
                                    print("Không có KEY trong danh sách.")
                                print("=================================\n")
                            else:
                                if len(lisss) == 7 or len(lisss) == 6 :
                                    print("Đây là PID ax+b")
                                    ws['N'+str(exc_var_2)] = "f(x)= a*x+b"
                                    a_write_can11_axcongb = ''
                                    b_write_can11_axcongb = ''
                                    Unitfx = ""
                                    keyunit = ""
                                    if len(lisss) == 7:
                                        print("a =   ", lisss[1].string)
                                        a_write_can11_axcongb = lisss[1].string
                                        ws['P'+str(exc_var_2)] = a_write_can11_axcongb
                                        b_write_can11_axcongb = lisss[2].string
                                        ws['Q'+str(exc_var_2)] = b_write_can11_axcongb
                                        ws['I'+str(exc_var_2)] = "N/A"
                                        ws['J'+str(exc_var_2)] = "N/A"
                                        ws['T'+str(exc_var_2)] = lisss[6].string
                                        

                                        if lisss[5].string != '_':
                                            Unitfx = lisss[5].string
                                            ws['U'+str(exc_var_2)] = Unitfx
                                            keyunit = lisss[5].string
                                            convertbb = 'Need Check'
                                            if keyunit in list_of_Unit:
                                                convertbb = Unit_sheets.loc[keyunit]
                                                ws['V'+str(exc_var_2)] = convertbb[0]
                                            else:
                                                ws['V'+str(exc_var_2)] = 'Need Check'
                                        else:
                                            Unitfx = 'None'
                                            ws['U'+str(exc_var_2)] = "N/A"
                                            keyunit = 'N/A'

                                        ws['R'+str(exc_var_2)] = lisss[4].string
                                        ws['S'+str(exc_var_2)] = lisss[3].string


                                    elif len(lisss) == 6: ##Case Freeform
                                        a_write_can11_axcongb = 'Need Check'
                                        ws['P'+str(exc_var_2)] = a_write_can11_axcongb
                                        b_write_can11_axcongb = lisss[1].string
                                        ws['Q'+str(exc_var_2)] = b_write_can11_axcongb
                                        if '>>' in lisss[1].string:
                                            hesosplit = lisss[1].string.split('>>')
                                            heso = hesosplit[1][0]
                                            ws['I'+str(exc_var_2)] = str(heso)
                                        elif '<<' in lisss[1].string:
                                            ws['I'+str(exc_var_2)] = "Need check bitwise case"
                                        else:
                                            ws['I'+str(exc_var_2)] = '0'
                                        
                                        ws['J'+str(exc_var_2)] = "Need Check: " + str(lisss[1].string)
                                        ws['T'+str(exc_var_2)] = lisss[5].string
                                        ws['U'+str(exc_var_2)] = lisss[4].string
                                        keyunit = 'N/A'
                                        Unitfx = 'None'
                                        ws['R'+str(exc_var_2)] = lisss[3].string
                                        ws['S'+str(exc_var_2)] = lisss[2].string
                                    
                                    ws['K'+str(exc_var_2)] = "N/A"
                                    ws['L'+str(exc_var_2)] = "N/A"
                                    ws['O'+str(exc_var_2)] = "N/A"                                                                                      
                                                                        
                                    Key_Check_GDS_Sp_FX = listmini[0].string + Unitfx                                                                                                        
                                    print("Key Check GDS Sp =   ", Key_Check_GDS_Sp_FX)
                                    ws['W'+str(exc_var_2)] = Key_Check_GDS_Sp_FX

                                    Key_axb = listmini[0].string + Command_AckCmd + Command_GetValueCmd + bytpos_write + listmini[2].string + 'N/AN/A' + a_write_can11_axcongb + b_write_can11_axcongb + keyunit
                                    print("Key ax+b =   ", Key_axb)
                                    ws['X'+str(exc_var_2)] = Key_axb

                                    Bien_kiemtra_GDS_1, Groupnamestring = support_GDS(Key_Check_GDS_Sp_FX, Sheet_GDS)
                                    if Bien_kiemtra_GDS_1 == True:
                                        ItemID = Mapping_a_x_plus_b(Key_axb, Sheet_ax_plus_b_Index_TableID)
                                    else:
                                        ItemID = 'GDS Not support'

                                    print("ItemID =   ", ItemID)
                                    ws['Z'+str(exc_var_2)] = ItemID
                                    ws['AA'+str(exc_var_2)] = Groupnamestring
                                    print("================================================\n")
                                else:
                                    print("Day la PID a&b")
                                    ws['N'+str(exc_var_2)] = "f(x)= x&a"
                                    print("Bit Position:  ", lisss[2].string)
                                    ws['I'+str(exc_var_2)] = lisss[2].string
                                    print("Bit Size:  ", lisss[3].string)
                                    ws['J'+str(exc_var_2)] = lisss[3].string
                                    ws['K'+str(exc_var_2)] = "N/A"
                                    ws['L'+str(exc_var_2)] = "N/A"
                                    print("Key Check GDS Sp =   ", listmini[0].string)
                                    ws['W'+str(exc_var_2)] = listmini[0].string + 'None'
                                    key_check_support_GDS = listmini[0].string + 'None'
                                    strrr = ""
                                    var_table = 1
                                    vitri_check = 5
                                    
                                    Default_text = 'DEFAULT:' + listmini[9].string
                                    if lisss[0].string == 'FreeFormSEV':
                                        vitri_check = 4
                                    for xx in range(vitri_check, len(lisss)):
                                        # print(lisss[xx].string)
                                        if var_table % 2 == 0:
                                            strrr = strrr + lisss[xx].string + "\n"
                                        else:
                                            strrr = strrr + lisss[xx].string + ":"
                                        var_table += 1
                                    print("---->>>:", strrr)
                                    # ws['O'+str(exc_var_2)] = strrr
                                    TableID_text = ""
                                    if len(strrr) == 0:
                                        ws['O'+str(exc_var_2)] = "No Table ID in DB"
                                        TableID_text = "No Table ID in DB"
                                    else:
                                        # strrr = strrr + Default_text + '\n'
                                        TableID_text, Table_Mising = split_Table(strrr, LIST_KEYYYYMAPPPING, Sheet_TableID, Sheet_TableID_Index_TableID)
                                        if TableID_text == 'No Table ID in DB':
                                            ID_Table_Missing = ID_Table_Missing + 1
                                            ID_Table_Missing_str = 'TableID_Miss_00' + str(ID_Table_Missing)
                                            Table_Mising.append(Default_text)
                                            Table_Mising.insert(0, ID_Table_Missing_str)
                                            for convertaa in convert_format(Table_Mising):
                                                list_Table_Missing.append(convertaa)
                                            ws['O'+str(exc_var_2)] = ID_Table_Missing_str
                                        else:                               
                                            ws['O'+str(exc_var_2)] = TableID_text
                                    
                                    
                                    try:
                                        if listmini[0].string is not None and listmini[1].string is not None and listmini[2].string is not None and lisss[2].string is not None and lisss[3].string is not None:
                                            # Chuyển đổi ký tự HTML entities thành dấu thích hợp
                                            if "&gt;" in lisss[2].string:
                                                lisss_2_value = lisss[2].string.replace("&gt;", ">")
                                                lisss_3_value = lisss[3].string.replace("&gt;", ">")
                                                Key_ab = listmini[0].string + Command_AckCmd + Command_GetValueCmd + bytpos_write + listmini[2].string + lisss_2_value + lisss_3_value + TableID_text
                                            else:
                                                lisss_2_value = lisss[2].string
                                                lisss_3_value = lisss[3].string                                      
                                                Key_ab = listmini[0].string + Command_AckCmd + Command_GetValueCmd + bytpos_write + listmini[2].string + lisss_2_value + lisss_3_value + TableID_text
                                        else:
                                            # Gán giá trị mặc định nếu có bất kỳ giá trị nào là None
                                            Key_ab = listmini[0].string + Command_AckCmd + Command_GetValueCmd + bytpos_write + listmini[2].string + TableID_text
                                    except TypeError as e:
                                        print(f"An exception occurred: {e}")
                                        # Xử lý ngoại lệ TypeError nếu cần thiết

                                    # print("Key_ab=   ", Key_ab)
                                    ws['Y'+str(exc_var_2)] = Key_ab

                                    Bien_kiemtra_GDS_2, Groupnamestring = support_GDS(key_check_support_GDS, Sheet_GDS)
                                    if Bien_kiemtra_GDS_2 == True:
                                        ItemID = Mapping_a_and_b(Key_ab, Sheet_a_and_b_Index_TableID)
                                    else:
                                        ItemID = 'GDS Not support'
                                    print("ItemID =   ", ItemID)

                                    ws['Z'+str(exc_var_2)] = ItemID
                                    ws['AA'+str(exc_var_2)] = Groupnamestring
                                    print("=========================\n")
                        elif CANTYPE == 'CAN29Bit':
                            print("Wait for me")

                            ws['I'+str(exc_var_2)] = listmini[1].string
                            ws['J'+str(exc_var_2)] = listmini[2].string

                            if len(listmini) > 10:
                                ReadConversion_Col = listmini[9]

                                Type = ReadConversion_Col.find('td')
                                if Type == None:
                                    Type_Col = listmini[9].string
                                else:
                                    Type_Col = Type.string
                                print('Your Type:', Type_Col)
                                Type_BaseEquation = False
                                if Type_Col == 'OdxLinear':
                                    ws['N'+str(exc_var_2)] = 'f(x)= a*x+b'
                                    Can29Unit = ""
                                    Can29float = ""
                                    Can29equation = ""
                                    Can29Min = ""
                                    Can29Max = ""
                                    can29_a = ""
                                    can29_b = ""
                                    if listmini[10].string != 'OdxLinear':
                                        Can29Unit = listmini[10].string
                                        Can29float = listmini[11].string
                                        Can29equation = listmini[12].string
                                        Can29Min = listmini[13].string
                                        Can29Max = listmini[15].string
                                        can29_a, can29_b = CAN29_equation(Can29equation)
                                        # if cHECK_BUG:
                                        #     raise TypeError
                                    else:
                                        Can29Unit = listmini[11].string
                                        Can29float = listmini[12].string
                                        Can29equation = listmini[13].string
                                        Can29Min = listmini[14].string
                                        Can29Max = listmini[16].string
                                        can29_a, can29_b = CAN29_equation(Can29equation)
                                        # if cHECK_BUG:
                                        #     raise TypeError
                                    ws['P'+str(exc_var_2)] = str(can29_a)
                                    ws['Q'+str(exc_var_2)] = str(can29_b)
                                    ws['R'+str(exc_var_2)] = Can29Min
                                    ws['S'+str(exc_var_2)] = Can29Max
                                    ws['T'+str(exc_var_2)] = Can29float

                                    Unitfx1 = ""
                                    keyunit1 = ""
                                    if Can29Unit != '_':
                                        Unitfx1 = Can29Unit
                                        ws['U'+str(exc_var_2)] = Unitfx1
                                        keyunit1 = Can29Unit
                                        convertbb1 = 'Need Check'
                                        if keyunit1 in list_of_Unit:
                                            convertbb1 = Unit_sheets.loc[keyunit1]
                                            ws['V'+str(exc_var_2)] = convertbb1[0]
                                        else:
                                            ws['V'+str(exc_var_2)] = 'Need Check'
                                        

                                    else:
                                        Unitfx1 = 'None'
                                        ws['U'+str(exc_var_2)] = "N/A"
                                        keyunit1 = 'N/A'
                                    
                                    #check GDS LD group items
                                    Key_Check_GDS_Sp_FX1 = listmini[0].string + Unitfx1
                                    print("Key Check GDS Sp =   ", Key_Check_GDS_Sp_FX1)
                                    ws['W'+str(exc_var_2)] = Key_Check_GDS_Sp_FX1


                                    Key_axb1 = listmini[0].string + Command_AckCmd  + Command_GetValueCmd + bytpos_write + bytesize_can29 + listmini[1].string +  listmini[2].string + str(can29_a) + str(can29_b) + keyunit1
                                    print("Key ax+b =   ", Key_axb1)
                                    ws['X'+str(exc_var_2)] = Key_axb1

                                    Bien_kiemtra_GDS_1, Groupnamestring = support_GDS(Key_Check_GDS_Sp_FX1, Sheet_GDS)
                                    if Bien_kiemtra_GDS_1 == True:
                                        ItemID = Mapping_a_x_plus_b(Key_axb1, Sheet_ax_plus_b_Index_TableID)
                                    else:
                                        ItemID = 'GDS Not support'

                                    print("ItemID =   ", ItemID)
                                    ws['Z'+str(exc_var_2)] = ItemID
                                    ws['AA'+str(exc_var_2)] = Groupnamestring
                                    print("================================================\n")
                                elif Type_Col == 'OdxTextTable' or Type_Col == 'FreeFormSEV':
                                    ws['N'+str(exc_var_2)] = 'f(x)= x&a'
                                    lisss = []
                                    if listmini[8].string == None and listmini[9].string == None:
                                        lisss = listmini[14:-1]
                                        if listmini[13].string == None:
                                            Type_BaseEquation = True
                                    else:
                                        lisss = listmini[13:-1]
                                        
                                    strrr = ""
                                    vitrix = 0
                                    vitritext = 4

                                    if Type_Col != 'FreeFormSEV':
                                        while vitritext < len(lisss):
                                            if lisss[vitrix].string == None:
                                                break
                                            else:
                                                strrr = strrr + lisss[vitrix].string + ":"
                                                strrr = strrr + lisss[vitritext].string + "\n"

                                                vitrix = vitritext + 1
                                                vitritext = vitrix + 4
                                    else:
                                        Check_case_have_OdxTextTable_FreeFormSEV = "NO_OK"
                                        if len(listmini) > 18:
                                            Check_case_have_OdxTextTable_FreeFormSEV = listmini[18].string
                                        if Check_case_have_OdxTextTable_FreeFormSEV == "NO_OK":
                                            var_table = 1
                                            for xx in range(0, len(lisss)):
                                                stringxxax = lisss[xx].string
                                                if stringxxax != None:
                                                    if var_table % 2 == 0:
                                                        strrr = strrr + stringxxax + "\n"
                                                    else:
                                                        strrr = strrr + stringxxax + ":"
                                                var_table += 1
                                        elif Check_case_have_OdxTextTable_FreeFormSEV ==  'OdxTextTable':
                                            vitrixa = 9
                                            vitritexta = 13
                                            while vitritexta < len(lisss):
                                                if lisss[vitrixa].string == None:
                                                    break
                                                else:
                                                    strrr = strrr + lisss[vitrixa].string + ":"
                                                    strrr = strrr + lisss[vitritexta].string + "\n"

                                                    vitrixa = vitritexta + 1
                                                    vitritexta = vitrixa + 4                                            
                                        else:
                                            if Type_BaseEquation:
                                                var_tabley = 1
                                                for yy in range(0, len(lisss)):
                                                    if var_tabley % 2 == 0:
                                                        strrr = strrr + lisss[yy].string + "\n"
                                                    else:
                                                        strrr = strrr + lisss[yy].string + ":"
                                                    var_tabley += 1


                                    print("---->>>:", strrr)
                                    # ws['O'+str(exc_var_2)] = strrr
                                    TableID_text = ""
                                    if len(strrr) == 0:
                                        ws['O'+str(exc_var_2)] = "No Table ID in DB"
                                        TableID_text = "No Table ID in DB"
                                    else:
                                        ## strrr = '0:Disabled\n1:Enabled\n'
                                        TableID_text, Table_Mising = split_Table(strrr, LIST_KEYYYYMAPPPING, Sheet_TableID, Sheet_TableID_Index_TableID)
                                        if TableID_text == 'No Table ID in DB':
                                            ID_Table_Missing = ID_Table_Missing + 1
                                            ID_Table_Missing_str = 'TableID_Miss_00' + str(ID_Table_Missing)
                                            Table_Mising.append('DEFAULT:UNRECOGNIZED STATE')
                                            Table_Mising.insert(0, ID_Table_Missing_str)
                                            for convertaa in convert_format(Table_Mising):
                                                list_Table_Missing.append(convertaa)
                                            TableID_text = ID_Table_Missing_str
                                            ws['O'+str(exc_var_2)] = TableID_text
                                        else:                               
                                            ws['O'+str(exc_var_2)] = TableID_text

                                    print('ok')

                                    ##Write_key 'f(x)= x&a'
                                    # key_a&b = ItemName + AckCmd + GetValueCmd + BytePosition +Bytesize + BitPosition + BitSize + TableID
                                    key_aandb = listmini[0].string + Command_AckCmd  + Command_GetValueCmd + bytpos_write + bytesize_can29 +  listmini[1].string +  listmini[2].string + TableID_text
                                    ws['Y'+str(exc_var_2)] = key_aandb


                                    key_check_support_GDS = listmini[0].string + 'None'
                                    ws['W'+str(exc_var_2)] = key_check_support_GDS

                                    Bien_kiemtra_GDS_2, Groupnamestring = support_GDS(key_check_support_GDS, Sheet_GDS)
                                    if Bien_kiemtra_GDS_2 == True:
                                        ItemID = Mapping_a_and_b(key_aandb, Sheet_a_and_b_Index_TableID)
                                    else:
                                        ItemID = 'GDS Not support'

                                    print("ItemID =   ", ItemID)

                                    ws['Z'+str(exc_var_2)] = ItemID
                                    ws['AA'+str(exc_var_2)] = Groupnamestring
                                    print("=========================\n")

                                elif Type_Col == 'OdxIdentical':
                                    ws['N'+str(exc_var_2)] = 'f(x)= HEX'
                                    key_check_support_GDS = listmini[0].string + 'None'
                                    ws['W'+str(exc_var_2)] = key_check_support_GDS
                                    ws['Z'+str(exc_var_2)] = 'Need Check'

                                elif Type_Col == 'Byte Array':
                                    ws['N'+str(exc_var_2)] = 'Byte Array'
                                    key_check_support_GDS = listmini[0].string + 'None'
                                    ws['W'+str(exc_var_2)] = key_check_support_GDS
                                    ws['Z'+str(exc_var_2)] = 'Need Check'
                                else:
                                    raise KeyError
                        exc_var_2 += 1
                    else:
                        continue
                # exc_var_2 += 1
            exc_var = exc_var_2
            print("**********************************************************************\n\n")
    savename = Excel_PathFile.replace('/ETI/', "/")
    savename = savename.replace('.html', '')
    savename = savename + '.xlsx'
    wb.save(savename)

    #DF chứa TableID missing:
    df_Table_Missing = pd.DataFrame(list_Table_Missing)
    #remove duplicate table
    New_DF_Table_Missing, DF_Replace  = check_Table_ID(df_Table_Missing)

    
    return savename, New_DF_Table_Missing, DF_Replace


# ################################
# # YOUR INPUT:
# # Your_Html_name = 'Audio Amplifier$GMNA_Audio Amplifier - Info 3 -  Global B.html'
# # Your_Html_name = 'Radio Controls$NGI - GMNA.html'
# Your_Html_name = '2021GMBuickEnclave.html'
# ###################################
# Excel_PathFile = 'DATA/2021GMBuickEnclave/ETI/' + Your_Html_name
# run_parser_html2excel(Excel_PathFile)

# expression = 'y=(1x+230)/10'
# a_return,b_return = find_coefficients(expression)
# print('a_return: ', a_return,'; b_return:', b_return)
